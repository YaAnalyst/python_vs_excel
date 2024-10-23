import openpyxl

data = openpyxl.load_workbook('online_retail.xlsx')
sheet = data.active

# Dictionary to store data by country
country_data = {}

# Iterate through the rows to collect data for each country
for row in range(2, sheet.max_row + 1):
    country = sheet.cell(row=row, column=8).value  # Country column (8th)
    if country not in country_data:
        country_data[country] = []
    
    customer_id = sheet.cell(row=row, column=7).value  # Customer ID (7th)
    quantity = sheet.cell(row=row, column=4).value  # Quantity (4th)
    
    # Append the relevant data (customer ID and quantity) for each country
    country_data[country].append((customer_id, quantity))

# Create separate worksheets for each country
for country, data_list in country_data.items():
    new_sheet = data.create_sheet(title=country)  # Create new sheet for each country
    new_sheet.cell(row=1, column=1, value="Customer ID")
    new_sheet.cell(row=1, column=2, value="Quantity")
    
    # Write the data into the new sheet
    for idx, (customer_id, quantity) in enumerate(data_list, start=2):
        new_sheet.cell(row=idx, column=1, value=customer_id)
        new_sheet.cell(row=idx, column=2, value=quantity)

# Save the updated workbook with new sheets
data.save('sales_data_by_country_2.xlsx')
